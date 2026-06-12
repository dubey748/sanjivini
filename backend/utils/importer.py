"""Excel/CSV import helpers for the admin portal.

Designed to be entity-agnostic where possible. Currently focused on
medicines (Phase 2). Returns rich per-row diagnostics for dry-run preview.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field, asdict
from typing import Any, Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

# ----- Medicine import schema -----------------------------------------------

MEDICINE_COLUMNS = [
    # (header, required, type, description)
    ("name",                True,  "str",   "Brand/Display name e.g. 'Paracetamol 500mg'"),
    ("brand",               False, "str",   "Brand name e.g. 'Crocin'"),
    ("composition",         True,  "str",   "Salt composition e.g. 'Paracetamol 500mg'"),
    ("category",            True,  "str",   "Category id (e.g. 'c-fever') OR category name"),
    ("price",               True,  "num",   "Selling price in INR"),
    ("mrp",                 False, "num",   "MRP. If blank, defaults to price"),
    ("pack",                False, "str",   "Pack description e.g. 'Strip of 15 tablets'"),
    ("prescription_required", False, "bool", "Yes/No — default No"),
    ("stock",               False, "int",   "On-hand stock units — default 0"),
    ("symptoms",            False, "str",   "Comma-separated symptoms for search"),
    ("manufacturer",        False, "str",   "Manufacturer name"),
    ("image_url",           False, "str",   "Single primary image URL"),
    ("image_urls",          False, "str",   "Semicolon-separated additional URLs"),
    ("sku",                 False, "str",   "Optional unique SKU (used for upsert matching)"),
    ("hsn_code",            False, "str",   "HSN/GST code"),
    ("gst_pct",             False, "num",   "GST percentage 0–28"),
    ("discount_pct",        False, "num",   "Default discount % 0–100"),
    ("tags",                False, "str",   "Semicolon-separated tags"),
    ("is_active",           False, "bool",  "Yes/No — default Yes"),
]

MEDICINE_HEADERS = [c[0] for c in MEDICINE_COLUMNS]
MEDICINE_REQUIRED = {c[0] for c in MEDICINE_COLUMNS if c[1]}

SAMPLE_ROWS = [
    {
        "name": "Paracetamol 500mg (Sample)", "brand": "Crocin", "composition": "Paracetamol 500mg",
        "category": "c-fever", "price": 28, "mrp": 35, "pack": "Strip of 15 tablets",
        "prescription_required": "No", "stock": 240, "symptoms": "fever, headache",
        "manufacturer": "GSK", "image_url": "https://images.pexels.com/photos/8666436/pexels-photo-8666436.jpeg",
        "image_urls": "", "sku": "SAMPLE-PARA-500", "hsn_code": "30049099", "gst_pct": 12,
        "discount_pct": 20, "tags": "fever;headache;otc", "is_active": "Yes",
    },
    {
        "name": "Azithromycin 500mg (Sample)", "brand": "Azithral", "composition": "Azithromycin 500mg",
        "category": "c-fever", "price": 96, "mrp": 120, "pack": "Strip of 5 tablets",
        "prescription_required": "Yes", "stock": 80, "symptoms": "bacterial infection",
        "manufacturer": "Alembic", "image_url": "https://images.pexels.com/photos/3683074/pexels-photo-3683074.jpeg",
        "image_urls": "", "sku": "SAMPLE-AZIT-500", "hsn_code": "30041040", "gst_pct": 5,
        "discount_pct": 20, "tags": "antibiotic;rx", "is_active": "Yes",
    },
]


# ----- Result containers ----------------------------------------------------

@dataclass
class RowResult:
    row_index: int                         # 1-based row number in source sheet (excluding header)
    action: str = "skip"                   # create | update | error | skip
    errors: list = field(default_factory=list)
    data: dict = field(default_factory=dict)
    match_id: Optional[str] = None         # existing medicine id when action='update'

    def is_ok(self) -> bool:
        return self.action in {"create", "update"} and not self.errors


@dataclass
class ImportResult:
    total: int = 0
    will_create: int = 0
    will_update: int = 0
    errors: int = 0
    rows: list = field(default_factory=list)

    def to_dict(self):
        d = asdict(self)
        return d


# ----- Helpers --------------------------------------------------------------

BOOL_TRUE = {"true", "yes", "y", "1", "on"}
BOOL_FALSE = {"false", "no", "n", "0", "off", ""}


def _normalize_header(h: Any) -> str:
    return re.sub(r"\s+", "_", str(h or "").strip().lower())


def _coerce(value: Any, kind: str):
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    if kind == "str":
        return s
    if kind == "num":
        return float(s.replace(",", ""))
    if kind == "int":
        return int(float(s.replace(",", "")))
    if kind == "bool":
        v = s.lower()
        if v in BOOL_TRUE: return True
        if v in BOOL_FALSE: return False
        raise ValueError(f"expected yes/no, got '{s}'")
    return s


def _parse_image_list(primary: Optional[str], extra: Optional[str]) -> list:
    out: list = []
    if primary:
        out.append(primary.strip())
    if extra:
        for token in re.split(r"[;\n]", extra):
            t = token.strip()
            if t and t not in out:
                out.append(t)
    return out


def _parse_tag_list(value: Optional[str]) -> list:
    if not value:
        return []
    return [t.strip() for t in re.split(r"[;\n]", value) if t.strip()]


# ----- Source readers -------------------------------------------------------

def _read_xlsx(file_bytes: bytes) -> tuple[list[str], Iterable[list]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], iter([])
    headers = [_normalize_header(h) for h in header_row]
    return headers, rows


def _read_csv(file_bytes: bytes) -> tuple[list[str], Iterable[list]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return [], iter([])
    headers = [_normalize_header(h) for h in header_row]
    return headers, reader


def read_table(filename: str, file_bytes: bytes) -> tuple[list[str], Iterable[list]]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _read_xlsx(file_bytes)
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _read_csv(file_bytes)
    # Best-effort detection
    if file_bytes[:2] == b"PK":
        return _read_xlsx(file_bytes)
    return _read_csv(file_bytes)


# ----- Medicine import logic -----------------------------------------------

async def _resolve_category(db, raw: str, category_index: dict) -> Optional[str]:
    """Accept category id OR name. Return canonical category id or None."""
    if not raw:
        return None
    key = str(raw).strip().lower()
    if key in category_index:
        return category_index[key]
    # Allow live lookup if the in-memory index missed (e.g. created mid-import)
    doc = await db.categories.find_one({"$or": [{"id": raw}, {"name": raw}]}, {"id": 1})
    if doc:
        return doc.get("id")
    return None


async def parse_medicine_table(
    db,
    filename: str,
    file_bytes: bytes,
    max_rows: int = 10000,
) -> ImportResult:
    headers, row_iter = read_table(filename, file_bytes)
    result = ImportResult()

    if not headers:
        result.errors = 1
        result.rows.append(RowResult(row_index=0, action="error", errors=["Empty or unreadable file"]))
        return result

    missing_required = MEDICINE_REQUIRED - set(headers)
    if missing_required:
        result.errors = 1
        result.rows.append(RowResult(
            row_index=0, action="error",
            errors=[f"Missing required column(s): {', '.join(sorted(missing_required))}"],
        ))
        return result

    column_kind = {c[0]: c[2] for c in MEDICINE_COLUMNS}
    col_idx = {h: headers.index(h) for h in headers}

    # Pre-load lookups
    cat_docs = await db.categories.find({}, {"id": 1, "name": 1}).to_list(500)
    category_index = {}
    for c in cat_docs:
        if c.get("id"):
            category_index[c["id"].lower()] = c["id"]
        if c.get("name"):
            category_index[c["name"].lower()] = c["id"]

    # Read all rows into list once, capping at max_rows
    raw_rows = []
    for r in row_iter:
        raw_rows.append(r)
        if len(raw_rows) >= max_rows:
            break

    # Build seen-key map within the file (catch dupes inside same upload)
    intra_seen: dict[str, int] = {}

    for i, raw in enumerate(raw_rows, start=1):
        rr = RowResult(row_index=i)
        row_dict = {h: (raw[col_idx[h]] if col_idx[h] < len(raw) else None) for h in headers}

        # Skip completely empty rows silently
        if all(v in (None, "") for v in row_dict.values()):
            rr.action = "skip"
            continue

        parsed: dict[str, Any] = {}
        for h in headers:
            if h not in column_kind:
                continue
            try:
                parsed[h] = _coerce(row_dict.get(h), column_kind[h])
            except Exception as e:
                rr.errors.append(f"{h}: {e}")

        # Required field check
        for req in MEDICINE_REQUIRED:
            if parsed.get(req) in (None, ""):
                rr.errors.append(f"{req}: required")

        # Category resolution
        if not rr.errors:
            resolved_cat = await _resolve_category(db, parsed.get("category"), category_index)
            if not resolved_cat:
                rr.errors.append(f"category: unknown '{parsed.get('category')}'")
            else:
                parsed["category"] = resolved_cat

        # Numeric sanity
        if "price" in parsed and parsed["price"] is not None:
            try:
                if float(parsed["price"]) < 0:
                    rr.errors.append("price: must be >= 0")
            except Exception:
                pass
        if parsed.get("mrp") in (None, "") and parsed.get("price") is not None:
            parsed["mrp"] = parsed["price"]
        if parsed.get("stock") in (None, ""):
            parsed["stock"] = 0
        if parsed.get("prescription_required") is None:
            parsed["prescription_required"] = False
        if parsed.get("is_active") is None:
            parsed["is_active"] = True

        # Image consolidation → `images` URL list, keep first as legacy `image`
        images = _parse_image_list(parsed.pop("image_url", None), parsed.pop("image_urls", None))
        parsed["images"] = images
        parsed["image"] = images[0] if images else None

        parsed["tags"] = _parse_tag_list(parsed.pop("tags", None))

        if rr.errors:
            rr.action = "error"
            rr.data = parsed
            result.rows.append(rr)
            continue

        # Duplicate detection
        match = None
        match_key = None
        sku = parsed.get("sku")
        if sku:
            match = await db.medicines.find_one({"sku": sku}, {"id": 1, "name": 1})
            match_key = f"sku:{sku.lower()}"
        if not match:
            match = await db.medicines.find_one({
                "name": {"$regex": f"^{re.escape(parsed['name'])}$", "$options": "i"},
                "composition": {"$regex": f"^{re.escape(parsed['composition'])}$", "$options": "i"},
                "pack": parsed.get("pack"),
            }, {"id": 1, "name": 1})
            match_key = f"name:{parsed['name'].lower()}|{parsed['composition'].lower()}|{(parsed.get('pack') or '').lower()}"

        # Detect intra-file duplicates (same key earlier in this sheet)
        if match_key in intra_seen:
            rr.errors.append(f"Duplicate of row {intra_seen[match_key]} in this file")
            rr.action = "error"
            rr.data = parsed
            result.rows.append(rr)
            continue
        intra_seen[match_key] = i

        if match:
            rr.action = "update"
            rr.match_id = match.get("id")
        else:
            rr.action = "create"
        rr.data = parsed
        result.rows.append(rr)

    # Totals
    result.total = len(result.rows)
    result.will_create = sum(1 for r in result.rows if r.action == "create")
    result.will_update = sum(1 for r in result.rows if r.action == "update")
    result.errors = sum(1 for r in result.rows if r.action == "error")
    return result


# ----- Template & export ----------------------------------------------------

def build_medicine_template_xlsx() -> bytes:
    """Return bytes of an .xlsx file with headers + 2 sample rows + Notes sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Medicines"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F4C3A")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(MEDICINE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.freeze_panes = "A2"

    # Column widths
    width_map = {"name": 28, "composition": 30, "image_url": 50, "image_urls": 50, "symptoms": 26, "tags": 22, "sku": 16}
    for col_idx, header in enumerate(MEDICINE_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width_map.get(header, 14)

    # Sample rows
    for r_idx, sample in enumerate(SAMPLE_ROWS, start=2):
        for c_idx, header in enumerate(MEDICINE_HEADERS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=sample.get(header, ""))

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Medicine Import — Field Guide"
    notes["A1"].font = Font(bold=True, size=14, color="0F4C3A")
    notes.append([])
    notes.append(["Column", "Required", "Type", "Notes"])
    for cell in notes[3]:
        cell.font = Font(bold=True)
    for c in MEDICINE_COLUMNS:
        notes.append([c[0], "Yes" if c[1] else "No", c[2], c[3]])
    notes.append([])
    notes.append(["Tips"])
    notes.append(["• 'category' accepts either the id (c-fever) or the visible name (Fever & Pain)."])
    notes.append(["• 'prescription_required' and 'is_active' accept Yes/No, true/false, 1/0."])
    notes.append(["• 'image_urls' is semicolon-separated. The 'image_url' column sets the primary."])
    notes.append(["• If 'sku' is provided it is used to match existing medicines for updates."])
    notes.append(["• Otherwise duplicates are detected by name + composition + pack (case-insensitive)."])
    notes.append(["• Up to 10,000 rows per upload. Split larger catalogs across files."])
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 10
    notes.column_dimensions["C"].width = 10
    notes.column_dimensions["D"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_medicines_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Medicines"
    for c_idx, h in enumerate(MEDICINE_HEADERS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F4C3A")
    ws.freeze_panes = "A2"
    for r_idx, doc in enumerate(rows, start=2):
        images = doc.get("images") or ([doc["image"]] if doc.get("image") else [])
        flat = {
            "name": doc.get("name"),
            "brand": doc.get("brand"),
            "composition": doc.get("composition"),
            "category": doc.get("category"),
            "price": doc.get("price"),
            "mrp": doc.get("mrp"),
            "pack": doc.get("pack"),
            "prescription_required": "Yes" if doc.get("prescription_required") else "No",
            "stock": doc.get("stock"),
            "symptoms": doc.get("symptoms"),
            "manufacturer": doc.get("manufacturer"),
            "image_url": images[0] if images else "",
            "image_urls": ";".join(images[1:]) if len(images) > 1 else "",
            "sku": doc.get("sku"),
            "hsn_code": doc.get("hsn_code"),
            "gst_pct": doc.get("gst_pct"),
            "discount_pct": doc.get("discount_pct"),
            "tags": ";".join(doc.get("tags") or []),
            "is_active": "Yes" if doc.get("is_active", True) else "No",
        }
        for c_idx, h in enumerate(MEDICINE_HEADERS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=flat.get(h))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_medicines_csv(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=MEDICINE_HEADERS)
    writer.writeheader()
    for doc in rows:
        images = doc.get("images") or ([doc["image"]] if doc.get("image") else [])
        writer.writerow({
            "name": doc.get("name", ""),
            "brand": doc.get("brand", ""),
            "composition": doc.get("composition", ""),
            "category": doc.get("category", ""),
            "price": doc.get("price", ""),
            "mrp": doc.get("mrp", ""),
            "pack": doc.get("pack", ""),
            "prescription_required": "Yes" if doc.get("prescription_required") else "No",
            "stock": doc.get("stock", 0),
            "symptoms": doc.get("symptoms", ""),
            "manufacturer": doc.get("manufacturer", ""),
            "image_url": images[0] if images else "",
            "image_urls": ";".join(images[1:]) if len(images) > 1 else "",
            "sku": doc.get("sku", "") or "",
            "hsn_code": doc.get("hsn_code", "") or "",
            "gst_pct": doc.get("gst_pct", "") or "",
            "discount_pct": doc.get("discount_pct", 0),
            "tags": ";".join(doc.get("tags") or []),
            "is_active": "Yes" if doc.get("is_active", True) else "No",
        })
    return buf.getvalue().encode("utf-8")
